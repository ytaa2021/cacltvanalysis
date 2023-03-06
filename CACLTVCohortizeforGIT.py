import csv
import functools
import openpyxl
import datetime

file = r"C:\Users\yotam.twersky\Downloads\accellionportaldb_pueports_churnwarehouse.xlsx"

#Churn Warehouse Data Loaded

wb = openpyxl.load_workbook(file)
ws = wb.active
cell_range = 'A2:F39694'
data = [[cell.value for cell in row] for row in ws[cell_range]]

#Structure: [ID, Expire Date, Platform, Metric, Amount, Account ID]

trans_file = r"C:\Users\yotam.twersky\OneDrive - Kiteworks\Documents\V8customer_account.xlsx"

#Customer Account Data Loaded

wb = openpyxl.load_workbook(trans_file)
ws = wb.active
cell_range = 'A1:B9069'
data2 = [[cell.value for cell in row] for row in ws[cell_range]]

#Structure: [id, sfid]

#SFIDs appended onto corresponding items of cw by account id
exceptions = []
for i in range(len(data)):
    try:
        data[i].append(list(filter(lambda j: j[0] == data[i][5], data2))[0][1])
    except:
        exceptions.append(data[i][5])

#prints a list of exception account ids (to take care of later)
#print(exceptions)

#Structure: [ID, Expire Date, Platform, Metric, Amount, Account ID, SFID]

cs_file = r"C:\Users\yotam.twersky\Downloads\TerritorryAccounts.xlsx"

#Customer Since File Loaded

wb = openpyxl.load_workbook(cs_file)
ws = wb.active
cell_range = 'A1:D21620'
data3 = [[cell.value for cell in row] for row in ws[cell_range]]

#[Account Name, Location, CS, SFID]



data3 = list(filter(lambda f: len(f) == 4, data3))

#Appends Customer Since, Account Name, and Location to the end of cw data

for i in range(len(data)):
    try:
        f1 = list(filter(lambda j: j[3] == str(data[i][6]), data3))
        if len(f1) == 1:
            data[i].append(f1[0][2])
            data[i].append(f1[0][0])
            data[i].append(f1[0][1])
    except:
        pass

print(data[0:3])
#Structure: [ID, Expire Date, Platform, Metric, Amount, Account ID, SFID, Customer Since, Account Name, Location]

#---------------------------------------------------------------------------------------------------------------------

#Combines Amount values and combines any rows with equal Expire_Date, Platform, Metric, and Account IDs. 
# (NOT USED)

'''
print(data[:5])
count = 0
data_new = []
print(len(data))
for i in data:
    filtered = list(filter(lambda x: x[1] == i[1] and x[2] == i[2] and x[3] == i[3] and x[5] == i[5], data))
    count = i[4]
    for f in range(len(filtered)-1):
        count += filtered[f+1][4]
        data.remove(filtered[f+1])
    try:
        data_new.append([i[0], i[1], i[2], i[3], count, i[5], i[6], i[7], i[8], i[9]])
    except:
        pass
'''

#to preserve the potential use of the above function in later versions
print("copying data")
data_new = data
#for i in data_new:
 #   i[:0] = [-1]

#Function to detect if a specific account ID's highest amount value falls within a certain minima and maxima
#Does so by finding the highest value abd then filtering that list of one value by the min/max specs and then 
#checks if that filtered list contains any items.
print("pre-progress")
key_list = list(set([i[5] for i in data_new]))
my_dict = {key: None for key in key_list}

for key in my_dict:
    if my_dict[key] == None:
        found_accounts = list(filter(lambda n: n[5] == key, data_new))
        if len(found_accounts) == 0:
            pass
        else:
            lowest = [0,datetime.datetime(2025, 5, 30, 0, 0),0,0,-1]
            for i in found_accounts:
                try:
                    if i[1] < lowest[1] and i[3] == "Expected":
                        lowest = i
                except:
                    pass
            if lowest == [0,datetime.datetime(2025, 5, 30, 0, 0)]:
                for i in found_accounts:  
                    if i[1] < lowest[1]:
                        lowest = i
            for i in found_accounts:
                if i[3] == "New Logo":
                    lowest = i
            my_dict[key] = lowest[4]

print(my_dict)
print("progress")

'''
This code should split into 7 tiers where 6 are based on $ values but for privacy I am hiding the tier definitions.

def check_cohort(row):
    try:
        if row[0] == 'KB':
            return 'Tier KB'
        if row[1] > num and row[1] < num:
            return 'Tier 1'
        if row[1] > num and row[1] < num:
            return 'Tier 2'
        if row[1] > num and row[1] < num:
            return 'Tier 3'
        if row[1] > num and row[1] < num:
            return 'Tier 4'
        if row[1] > num and row[1] < num:
            return 'Tier 5'
        if row[1] > num and row[1] < num:
            return 'Tier 6'
        return 'Tier None'
    except:
        return 'Tier None'

'''

for i in data_new:
    i[:0] = [check_cohort([i[2],my_dict[i[5]]])]



#calls on the check_cohort() function to append the correct cohort by highest transaction 
#to the end of each item in data_new

#Exports the list data_new to an already specified csv file

doofile = open(r"C:\Users\yotam.twersky\OneDrive - Kiteworks\Documents\churn8.csv", 'w')
csv_writer = csv.writer(doofile)
csv_writer.writerows(data_new)
print('done')

